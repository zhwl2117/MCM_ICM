from openpyxl import load_workbook
from most_popular import get_10_most_popular
from most_popular import read_excel
from WIlson import *
import xlwt
import matplotlib.pyplot as plt


def xls_select_star_year(xls_str, key, year):
    wb = load_workbook(xls_str)
    ws = wb[wb.sheetnames[0]]
    message = []
    for row in ws.rows:
        if not isinstance(row[3].value, int):
            if row[3].value.find(key) >= 0 and row[14].value.find(year) >= 0:
                message.append(row[7].value)
    return message


def xls_select_wilson(name):
    pa = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    wb = load_workbook(pa)
    ws = wb[wb.sheetnames[0]]
    message = []
    stat = 1
    for row in ws.rows:
        if stat == 1:
            stat += 1
            continue
        if not isinstance(row[3].value, int):
            print(row[13].value, type(row[13].value))
            if len(str(row[13].value).split()) >= 55:
                score = wilson_score(row[8].value, row[9].value)
                msg = []
                for i in range(len(row)):
                    msg.append(row[i].value)
                msg.append(score)
                message.append(msg)
    return message


def write_data(message):
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('My Worksheet')
    row = 0
    for msg in message:
        for i in range(len(msg)):
            worksheet.write(row, i, label=str(msg[i]))
        row += 1
    workbook.save(r'C:\Users\17998\Desktop\MCM_ICN\hawkes_pacifier.xls')


def xls_select_time(xls_str, key):
    wb = load_workbook(xls_str)
    ws = wb[wb.sheetnames[0]]
    message = []
    for row in ws.rows:
        if not isinstance(row[3].value, int):
            if row[3].value.find(key) >= 0:
                message.append([row[7].value, row[14].value])
    return message


def xls_select_star(xls_str, key):
    wb = load_workbook(xls_str)
    ws = wb[wb.sheetnames[0]]
    message = []
    for row in ws.rows:
        if not isinstance(row[3].value, int):
            if row[3].value.find(key) >= 0:
                message.append(row[7].value)
    return message


def xls_select_year(xls_str, key, year):
    wb = load_workbook(xls_str)
    ws = wb[wb.sheetnames[0]]
    message = []
    for row in ws.rows:
        if not isinstance(row[3].value, int):
            if row[3].value.find(key) >= 0 and row[14].value.find(year):
                message.append((row[8].value, row[9].value, row[15].value))
    return message


def xls_select(xls_str, key):
    wb = load_workbook(xls_str)
    ws = wb[wb.sheetnames[0]]
    message = []
    for row in ws.rows:
        if not isinstance(row[3].value, int):
            if row[3].value.find(key) >= 0:
                message.append((row[8].value, row[9].value, row[15].value))
    return message


def get_msg_year(name):
    path = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    years = range(9, 19)
    most = get_10_most_popular(read_excel(name))
    msgss = []
    for i in range(len(most)):
        msgs = []
        for year in years:
            msgs.append(xls_select_year(path, most[i][0], str(year)))
        msgss.append(msgs)
    return msgss


def get_msg_star_year(name):
    path = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    years = range(9, 19)
    most = get_10_most_popular(read_excel(name))
    msgss = []
    for i in range(len(most)):
        msgs = []
        for year in years:
            msgs.append(xls_select_star_year(path, most[i][0], str(year)))
        msgss.append(msgs)
    return msgss


def get_msg(name):
    path = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    most = get_10_most_popular(read_excel(name))
    msgs = []
    for i in range(len(most)):
        msgs.append(xls_select(path, most[i][0]))
    return msgs


def get_msg_star(name):
    path = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    most = get_10_most_popular(read_excel(name))
    msgs = []
    for i in range(len(most)):
        msgs.append(xls_select_star(path, most[i][0]))
    return msgs


if __name__ == "__main__":
    name = "pacifier"
    pat = r"C:\Users\17998\Desktop\MCM_ICN" + "\\" + name + ".xlsx"
    msgs = xls_select_time(pat, "B003CK3LDI")
    write_data(msgs)

